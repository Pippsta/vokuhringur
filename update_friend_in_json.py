"""Update only the friend records in observations.json from awake_times.xlsx.

This is the local refresh path — chess records in observations.json are owned
by the GitHub Action (which fetches them weekly from chess.com + lichess) and
should NOT be regenerated from local xlsx files (which are static copies that
fall behind the bot's fetch).

The script preserves all 'chess' records as-is and replaces all 'friend'
records with whatever is currently in awake_times.xlsx.

Used by refresh.cmd. For a full local rebuild from local xlsx (rare — only if
you've just re-exported chess.com / lichess data locally), run convert.py
directly instead.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
JSON_PATH = HERE / "observations.json"
XLSX_PATH = HERE / "awake_times.xlsx"


def main() -> int:
    if not JSON_PATH.exists():
        print(f"ERROR: {JSON_PATH.name} not found", file=sys.stderr)
        return 1
    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH.name} not found", file=sys.stderr)
        return 1

    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    # Preserve chess records — the GitHub Action owns these.
    chess = [o for o in data["observations"] if o.get("source") == "chess"]

    # Read fresh friend records from the local xlsx.
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)  # discard header
    friend: list[dict] = []
    for row in rows:
        ts = row[0]
        if ts is None:
            continue
        iso = ts.replace(tzinfo=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        friend.append({"ts": iso, "source": "friend"})

    merged = sorted(chess + friend, key=lambda r: r["ts"])
    data["observations"] = merged
    data["counts"]["by_source"] = {"chess": len(chess), "friend": len(friend)}
    data["counts"].setdefault("input_rows", {})["friend"] = len(friend)
    data["counts"]["total"] = len(merged)
    data["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    JSON_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(
        f"Updated friend records: chess={len(chess)} (preserved), "
        f"friend={len(friend)}, total={len(merged)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
