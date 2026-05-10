"""Fetch every published chess.com game for $CHESS_USERNAME and write to chess_games.xlsx.

Used by the GitHub Action (.github/workflows/refresh.yml). Configurable via env vars:

  CHESS_USERNAME — required, the chess.com handle to fetch
  CONTACT_EMAIL  — recommended, sent in the User-Agent header (chess.com requires
                   a contact in User-Agent for archive endpoints)

Output is written to chess_games.xlsx next to this script.
"""

from __future__ import annotations

import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

USERNAME = os.environ.get("CHESS_USERNAME")
if not USERNAME:
    sys.exit("ERROR: CHESS_USERNAME env var is required")

CONTACT_EMAIL = os.environ.get("CONTACT_EMAIL", "")
USER_AGENT = f"chess-export/1.0 (contact: {CONTACT_EMAIL})"

HERE = Path(__file__).parent
OUTPUT_FILE = HERE / "chess_games.xlsx"
BASE_URL = "https://api.chess.com/pub"
REQUEST_SLEEP = 0.25
MAX_RETRIES = 6

DRAW_RESULTS = {
    "agreed", "repetition", "stalemate", "insufficient",
    "50move", "timevsinsufficient",
}
WIN_RESULTS = {"win"}

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT, "Accept": "application/json"})


def get_json(url: str) -> dict:
    """GET with polite sleep + backoff on connection errors, 429, and 5xx."""
    for attempt in range(MAX_RETRIES):
        try:
            resp = session.get(url, timeout=30)
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            wait = 2 ** attempt
            print(f"  Connection error on {url}: {exc.__class__.__name__} — sleeping {wait:.1f}s (attempt {attempt + 1}/{MAX_RETRIES})")
            time.sleep(wait)
            continue
        if resp.status_code == 200:
            time.sleep(REQUEST_SLEEP)
            return resp.json()
        if resp.status_code == 429 or 500 <= resp.status_code < 600:
            wait = float(resp.headers.get("Retry-After", 0)) or (2 ** attempt)
            print(f"  HTTP {resp.status_code} on {url} — sleeping {wait:.1f}s (attempt {attempt + 1}/{MAX_RETRIES})")
            time.sleep(wait)
            continue
        resp.raise_for_status()
    raise RuntimeError(f"Exceeded retries for {url}")


def normalize_outcome(my_result: str) -> str:
    if my_result in WIN_RESULTS:
        return "win"
    if my_result in DRAW_RESULTS:
        return "draw"
    return "loss"


def extract_row(game: dict) -> dict:
    white = game.get("white", {}) or {}
    black = game.get("black", {}) or {}
    white_user = (white.get("username") or "").strip()
    black_user = (black.get("username") or "").strip()

    if white_user.lower() == USERNAME.lower():
        played_as = "white"
        my_side, opp_side = white, black
        opponent = black_user
    else:
        played_as = "black"
        my_side, opp_side = black, white
        opponent = white_user

    my_result = my_side.get("result", "") or ""
    my_rating = my_side.get("rating")
    opp_rating = opp_side.get("rating")
    rating_diff = (
        my_rating - opp_rating
        if isinstance(my_rating, (int, float)) and isinstance(opp_rating, (int, float))
        else None
    )

    end_ts = game.get("end_time")
    end_dt = datetime.fromtimestamp(end_ts, tz=timezone.utc) if end_ts else None
    end_naive = end_dt.replace(tzinfo=None) if end_dt else None
    date_only = end_dt.date() if end_dt else None
    day_of_week = end_dt.strftime("%A") if end_dt else None
    hour = end_dt.hour if end_dt else None

    return {
        "EndTime (UTC)": end_naive,
        "Date": date_only,
        "DayOfWeek": day_of_week,
        "Hour": hour,
        "TimeClass": game.get("time_class"),
        "TimeControl": game.get("time_control"),
        "Rated": game.get("rated"),
        "Rules": game.get("rules"),
        "PlayedAs": played_as,
        "Opponent": opponent,
        "MyResult": my_result,
        "Outcome": normalize_outcome(my_result),
        "MyRating": my_rating,
        "OpponentRating": opp_rating,
        "RatingDiff": rating_diff,
        "WhiteUser": white_user,
        "WhiteRating": white.get("rating"),
        "WhiteResult": white.get("result"),
        "BlackUser": black_user,
        "BlackRating": black.get("rating"),
        "BlackResult": black.get("result"),
        "GameUrl": game.get("url"),
        "_sort_key": end_ts or 0,
    }


def fetch_all_games() -> list[dict]:
    archives_url = f"{BASE_URL}/player/{USERNAME}/games/archives"
    print(f"Fetching archive list: {archives_url}")
    archives = get_json(archives_url).get("archives", [])
    print(f"  Found {len(archives)} monthly archives")

    rows: list[dict] = []
    for i, url in enumerate(archives, 1):
        print(f"[{i}/{len(archives)}] {url}")
        data = get_json(url)
        games = data.get("games", []) or []
        for g in games:
            rows.append(extract_row(g))
        print(f"    +{len(games)} games (running total: {len(rows)})")

    rows.sort(key=lambda r: r["_sort_key"], reverse=True)
    for r in rows:
        r.pop("_sort_key", None)
    return rows


def write_xlsx(rows: list[dict], path: Path) -> None:
    columns = [
        ("EndTime (UTC)", 19, "yyyy-mm-dd hh:mm:ss"),
        ("Date", 12, "yyyy-mm-dd"),
        ("DayOfWeek", 11, None),
        ("Hour", 6, None),
        ("TimeClass", 11, None),
        ("TimeControl", 12, None),
        ("Rated", 7, None),
        ("Rules", 10, None),
        ("PlayedAs", 9, None),
        ("Opponent", 22, None),
        ("MyResult", 16, None),
        ("Outcome", 9, None),
        ("MyRating", 10, None),
        ("OpponentRating", 14, None),
        ("RatingDiff", 11, None),
        ("WhiteUser", 22, None),
        ("WhiteRating", 12, None),
        ("WhiteResult", 16, None),
        ("BlackUser", 22, None),
        ("BlackRating", 12, None),
        ("BlackResult", 16, None),
        ("GameUrl", 60, None),
    ]
    headers = [c[0] for c in columns]

    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    arial = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="FF1F2A44", end_color="FF1F2A44")
    header_align = Alignment(horizontal="left", vertical="center")

    ws.append(headers)
    for col_idx, (_, width, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r in rows:
        ws.append([r.get(h) for h in headers])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx, (_, _, num_fmt) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = arial
            if num_fmt:
                cell.number_format = num_fmt

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.row_dimensions[1].height = 22

    wb.save(path)


def main() -> None:
    rows = fetch_all_games()
    write_xlsx(rows, OUTPUT_FILE)
    print()
    print(f"Total games fetched: {len(rows)}")
    print(f"Output file:         {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        sys.exit(130)
