"""Reconstruct awake_times.xlsx from the committed observations.json.

Used inside the GitHub Action runner only. The local `awake_times.xlsx` is
gitignored — it lives only on the user's machine and OneDrive — so the runner
doesn't have it. This script extracts friend records from observations.json and
writes them back into the single-column xlsx convert.py expects.

Soft-fails on missing observations.json (writes empty xlsx and continues —
useful for first-ever run on a fresh clone). Hard-fails on malformed JSON or
missing 'observations' key — protects friend records from being silently
overwritten by a corrupt JSON.
"""

from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
JSON_INPUT = HERE / "observations.json"
XLSX_OUTPUT = HERE / "awake_times.xlsx"


def main() -> int:
    ts_list: list[datetime] = []

    if JSON_INPUT.exists():
        try:
            data = json.loads(JSON_INPUT.read_text(encoding="utf-8"))
        except json.JSONDecodeError as e:
            print(f"ERROR: observations.json invalid: {e}", file=sys.stderr)
            return 1
        observations = data.get("observations")
        if not isinstance(observations, list):
            print("ERROR: observations.json missing 'observations' list", file=sys.stderr)
            return 1
        for r in observations:
            if r.get("source") != "friend":
                continue
            ts = r.get("ts")
            if not isinstance(ts, str):
                continue
            try:
                # Match _iso_utc's format exactly so the round-trip is byte-identical.
                ts_list.append(datetime.strptime(ts, "%Y-%m-%dT%H:%M:%SZ"))
            except ValueError:
                print(f"WARN: skipping unparseable friend ts: {ts!r}", file=sys.stderr)
    else:
        print(f"observations.json not found — writing empty {XLSX_OUTPUT.name}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Awake"

    # Header styling matches the local awake_times.xlsx so a downloaded artifact
    # is human-inspectable.
    arial = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="FF1F2A44", end_color="FF1F2A44")
    header_align = Alignment(horizontal="left", vertical="center")

    ws.append(["Awake Timestamp"])
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = header_align
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.row_dimensions[1].height = 22

    # Sort ascending so the file matches the convention of the local one
    # (which the user appends to in chronological order).
    for dt in sorted(ts_list):
        ws.append([dt])

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = arial
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = "A2"

    wb.save(XLSX_OUTPUT)
    print(f"Wrote {XLSX_OUTPUT.name}: {len(ts_list)} friend records")
    return 0


if __name__ == "__main__":
    sys.exit(main())
