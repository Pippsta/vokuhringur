#!/usr/bin/env python3
"""
convert.py — chess + friend xlsx -> observations.json

Reads two awake-evidence sources and emits the merged JSON the static page
consumes:

  - eidurm_games.xlsx       (chess.com endgame timestamps, machine-recorded)
  - friend_awake_times.xlsx (friend observations, manual, minute precision)

Drops daily/correspondence chess games — endpoint is too weak as awake-evidence
(the final move can be hours or days after the player was actually awake).

Friend timestamps have no tzinfo; treated as UTC because Iceland is UTC
year-round and observations are recorded in Iceland local time.

See README.md for the full refresh process.
"""

import json
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent
CHESS_INPUT = HERE / "eidurm_games.xlsx"
FRIEND_INPUT = HERE / "friend_awake_times.xlsx"
OUTPUT = HERE / "observations.json"
SUBJECT = "Edson"
EXCLUDED_TIME_CLASSES = {"daily"}


def _iso_utc(dt: datetime) -> str:
    return dt.replace(tzinfo=timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def read_chess(path: Path) -> tuple[list[dict], int, int]:
    """Return (records, input_rows, excluded_daily_games)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Games"]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    end_idx = header.index("EndTime (UTC)")
    class_idx = header.index("TimeClass")

    records: list[dict] = []
    excluded_daily = 0
    input_rows = 0
    for row in rows:
        input_rows += 1
        ts = row[end_idx]
        time_class = row[class_idx]
        if ts is None:
            continue
        if time_class in EXCLUDED_TIME_CLASSES:
            excluded_daily += 1
            continue
        records.append({"ts": _iso_utc(ts), "source": "chess"})
    return records, input_rows, excluded_daily


def read_friend(path: Path) -> tuple[list[dict], int]:
    """Return (records, input_rows)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)  # discard header

    records: list[dict] = []
    input_rows = 0
    for row in rows:
        input_rows += 1
        ts = row[0]
        if ts is None:
            continue
        records.append({"ts": _iso_utc(ts), "source": "friend"})
    return records, input_rows


def main() -> int:
    missing = [p.name for p in (CHESS_INPUT, FRIEND_INPUT) if not p.exists()]
    if missing:
        print(f"ERROR: missing input file(s): {', '.join(missing)}", file=sys.stderr)
        return 1

    chess_recs, chess_in, excluded_daily = read_chess(CHESS_INPUT)
    friend_recs, friend_in = read_friend(FRIEND_INPUT)

    merged = chess_recs + friend_recs
    merged.sort(key=lambda r: r["ts"])

    payload = {
        "subject": SUBJECT,
        "sources": {
            "chess": "chess.com endgame timestamps (machine-recorded, session-clustered, second-precision)",
            "friend": "friend observations (manual, minute-precision, point observations)",
        },
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "timezone_note": (
            "Iceland is UTC year-round (no DST); ts is UTC and equals local time. "
            "Friend timestamps are entered in Iceland local time and treated as UTC."
        ),
        "counts": {
            "by_source": {"chess": len(chess_recs), "friend": len(friend_recs)},
            "input_rows": {"chess": chess_in, "friend": friend_in},
            "excluded_daily_games": excluded_daily,
            "total": len(merged),
        },
        "observations": merged,
    }

    OUTPUT.write_text(
        json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(
        f"Wrote {OUTPUT.name}: "
        f"chess={len(chess_recs)} (excluded_daily={excluded_daily}), "
        f"friend={len(friend_recs)}, total={len(merged)}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
