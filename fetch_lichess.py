"""Fetch every Lichess game for $LICHESS_USERNAME and write timestamps to lichess_games.xlsx.

Used by the GitHub Action (.github/workflows/refresh.yml). Configurable via env vars:

  LICHESS_USERNAME — required, the lichess handle to fetch
  CONTACT_EMAIL    — recommended, sent in the User-Agent header (lichess prefers it)

Output is written to lichess_games.xlsx next to this script.
"""

from __future__ import annotations

import json
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

USERNAME = os.environ.get("LICHESS_USERNAME")
if not USERNAME:
    sys.exit("ERROR: LICHESS_USERNAME env var is required")

CONTACT_EMAIL = os.environ.get("CONTACT_EMAIL", "")
USER_AGENT = f"lichess-export/1.0 (contact: {CONTACT_EMAIL})"

HERE = Path(__file__).parent
OUTPUT_FILE = HERE / "lichess_games.xlsx"
URL = f"https://lichess.org/api/games/user/{USERNAME}"
PARAMS = {
    "pgnInJson": "false",
    "clocks": "false",
    "evals": "false",
    "opening": "false",
    "moves": "false",
    "tags": "false",
}
MAX_RETRIES = 6


def fetch_end_times() -> list[datetime]:
    headers = {"User-Agent": USER_AGENT, "Accept": "application/x-ndjson"}
    end_times: list[datetime] = []

    for attempt in range(MAX_RETRIES):
        print(f"GET {URL} (attempt {attempt + 1}/{MAX_RETRIES})")
        try:
            resp_cm = requests.get(URL, params=PARAMS, headers=headers, stream=True, timeout=120)
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            wait = 2 ** attempt
            print(f"  Connection error: {exc.__class__.__name__} — sleeping {wait:.1f}s")
            time.sleep(wait)
            continue
        with resp_cm as resp:
            if resp.status_code == 429:
                wait = float(resp.headers.get("Retry-After", 0)) or (2 ** attempt)
                print(f"  HTTP 429 — sleeping {wait:.1f}s")
                time.sleep(wait)
                continue
            if 500 <= resp.status_code < 600:
                wait = 2 ** attempt
                print(f"  HTTP {resp.status_code} — sleeping {wait:.1f}s")
                time.sleep(wait)
                continue
            resp.raise_for_status()

            count = 0
            for line in resp.iter_lines(decode_unicode=True):
                if not line:
                    continue
                game = json.loads(line)
                ts_ms = game.get("lastMoveAt") or game.get("createdAt")
                if ts_ms is None:
                    continue
                dt = datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).replace(tzinfo=None)
                end_times.append(dt)
                count += 1
                if count % 200 == 0:
                    print(f"  ...{count} games parsed")
            print(f"  Done: {count} games parsed")
            return end_times

    raise RuntimeError("Exceeded retries")


def write_xlsx(end_times: list[datetime], path: Path) -> None:
    end_times_sorted = sorted(end_times, reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    arial = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="FF1F2A44", end_color="FF1F2A44")
    header_align = Alignment(horizontal="left", vertical="center")

    ws.append(["EndTime (UTC)"])
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = header_align
    ws.column_dimensions[get_column_letter(1)].width = 22
    ws.row_dimensions[1].height = 22

    for dt in end_times_sorted:
        ws.append([dt])

    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1)
        cell.font = arial
        cell.number_format = "yyyy-mm-dd hh:mm:ss"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:A{ws.max_row}"

    wb.save(path)


def main() -> None:
    end_times = fetch_end_times()
    write_xlsx(end_times, OUTPUT_FILE)
    print()
    print(f"Total games fetched: {len(end_times)}")
    print(f"Output file:         {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted.", file=sys.stderr)
        sys.exit(130)
